import tkinter as tk
from tkinter import messagebox
from tkinter import font
from openpyxl import Workbook, load_workbook
import os

def submit_data():
    data = {
        'Name': entry_name.get(),
        'Email': entry_email.get(),
        'Contact No': entry_contact.get(),
        'Address': entry_address.get()
    }

    # Check if Excel file exists
    if not os.path.exists('data.xlsx'):
        # Create a new workbook and add headings
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Contact No', 'Address'])
    else:
        # Load existing workbook
        wb = load_workbook('data.xlsx')
        ws = wb.active
    
    # Append data to the worksheet
    ws.append([data['Name'], data['Email'], data['Contact No'], data['Address']])
    
    # Save the workbook
    wb.save('data.xlsx')
    
    messagebox.showinfo('Success', 'Data saved successfully!')


def center_window(root, width=300, height=200):
    # Get screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calculate position x, y
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    root.geometry('%dx%d+%d+%d' % (width, height, x, y))

# Creating main window
root = tk.Tk()
root.title("Registration Form")

# Set window size and position
center_window(root, 400, 250)

# Style settings
label_font = font.Font(size=12)
entry_font = font.Font(size=12)
header_font = font.Font(size=16, weight='bold')

# Heading label
tk.Label(root, text="Registration Form", font=header_font).grid(row=0, column=0, columnspan=2, pady=10)

# Create labels and entry fields
tk.Label(root, text="Name", font=label_font).grid(row=1, column=0, sticky=tk.W, padx=20, pady=5)
entry_name = tk.Entry(root, font=entry_font)
entry_name.grid(row=1, column=1, padx=20, pady=5)

tk.Label(root, text="Email", font=label_font).grid(row=2, column=0, sticky=tk.W, padx=20, pady=5)
entry_email = tk.Entry(root, font=entry_font)
entry_email.grid(row=2, column=1, padx=20, pady=5)

tk.Label(root, text="Contact No", font=label_font).grid(row=3, column=0, sticky=tk.W, padx=20, pady=5)
entry_contact = tk.Entry(root, font=entry_font)
entry_contact.grid(row=3, column=1, padx=20, pady=5)

tk.Label(root, text="Address", font=label_font).grid(row=4, column=0, sticky=tk.W, padx=20, pady=5)
entry_address = tk.Entry(root, font=entry_font)
entry_address.grid(row=4, column=1, padx=20, pady=5)

# Submit button
submit_button = tk.Button(root, text="Submit", command=submit_data, font=label_font, bg="#4CAF50", fg="white", padx=10, pady=5)
submit_button.grid(row=5, column=0, columnspan=2, pady=20)

root.mainloop()
